"""Reads the Excel file to extract exemplar ads and campaign specs."""

import openpyxl
from config import EXCEL_PATH, SHEET_MAP


def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)


def get_rsa_template(wb):
    """Get the RSA template structure from SGS_146 sheet."""
    ws = wb["SGS_146"]
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers.append({"col": col, "name": str(val).strip()})

    headline_cols = [h for h in headers if "Titular" in h["name"] or "Tiular" in h["name"]]
    desc_cols = [h for h in headers if "Descripción" in h["name"] or "Descipción" in h["name"]]

    return {
        "headline_count": len(headline_cols),
        "description_count": len(desc_cols),
        "headline_max_chars": 30,
        "description_max_chars": 90,
    }


def get_campaign_specs(wb, marca: str, pilar: str = None):
    """Extract campaign specs from the corresponding Ads Media sheet."""
    sheet_name = SHEET_MAP.get(marca)
    if not sheet_name or sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    campaigns = []

    for row_idx in range(2, ws.max_row + 1):
        ticket = ws.cell(row=row_idx, column=3).value
        if not ticket:
            continue

        channel = str(ws.cell(row=row_idx, column=4).value or "")
        pillar = str(ws.cell(row=row_idx, column=12).value or "")
        specs_copy = str(ws.cell(row=row_idx, column=11).value or "")
        general_req = str(ws.cell(row=row_idx, column=6).value or "")
        refs = str(ws.cell(row=row_idx, column=15).value or "")
        content_type = str(ws.cell(row=row_idx, column=7).value or "")
        market = str(ws.cell(row=row_idx, column=9).value or "")
        tech_specs = str(ws.cell(row=row_idx, column=10).value or "")

        if pilar:
            pilar_lower = pilar.lower()
            pillar_lower = pillar.lower()
            if pilar_lower == "caracteristicas" and "caracter" not in pillar_lower:
                continue
            if pilar_lower == "beneficios" and "benefic" not in pillar_lower:
                continue
            if pilar_lower == "marca" and pillar_lower != "marca":
                continue

        campaigns.append({
            "ticket": str(ticket),
            "channel": channel,
            "pillar": pillar,
            "specs_copy": specs_copy,
            "general_requirement": general_req,
            "references": refs,
            "content_type": content_type,
            "market": market,
            "tech_specs": tech_specs,
        })

    return campaigns


def get_exemplar_ads(wb, marca: str):
    """Extract existing ad copy examples from the sheet to use as exemplars."""
    campaigns = get_campaign_specs(wb, marca)
    exemplars = []

    for c in campaigns:
        if c["specs_copy"] and len(c["specs_copy"]) > 20:
            exemplars.append({
                "pillar": c["pillar"],
                "copy": c["specs_copy"][:300],
                "refs": c["references"][:300] if c["references"] != "N/A" else "",
            })

    return exemplars
