#!/usr/bin/env python3
"""Generador de copys para Skydropx — Google Ads y Meta Ads.

Usage:
    python generate_ads.py
"""

import json
import os
import re
import sys
from datetime import datetime

import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from brand_parser import build_style_guide_prompt, STYLE_GUIDE, PILAR_CONTENT
from config import (
    GEMINI_API_KEY, AD_SPECS, OUTPUT_DIR, CANAL_FORMATS,
    OBJETIVO_PROMPT, ENFOQUE_NARRATIVO_PROMPT, is_meta_format,
)
from exemplars import (
    get_headline_exemplars_for_prompt,
    get_description_exemplars_for_prompt,
    COPY_PATTERNS,
    CTA_EXEMPLARS,
    META_COPY_PATTERNS,
    META_CTA_EXEMPLARS,
    get_meta_post_copy_exemplars_for_prompt,
    get_meta_encabezado_exemplars_for_prompt,
    get_meta_descripcion_exemplars_for_prompt,
    get_meta_copy_imagen_exemplars_for_prompt,
    get_meta_video_exemplars_for_prompt,
)
from validator import validate_ad_set, print_validation_report
from learning import get_learning_prompt_section


def configure_gemini():
    """Configure Gemini API. Returns False if key is missing (for Streamlit)."""
    if not GEMINI_API_KEY:
        try:
            import streamlit as _st
            _st.error(
                "GEMINI_API_KEY no configurada. "
                "Agrega `GEMINI_API_KEY = \"tu-key\"` en Settings > Secrets."
            )
            _st.stop()
        except ImportError:
            print("Error: GEMINI_API_KEY no configurada.")
            print("Ejecuta: export GEMINI_API_KEY='tu-api-key'")
            sys.exit(1)
    genai.configure(api_key=GEMINI_API_KEY)


# ---------------------------------------------------------------------------
# Interactive wizard
# ---------------------------------------------------------------------------

def _ask_choice(prompt_text, options, default=None):
    """Ask user to pick from a list of options."""
    print(f"\n  {prompt_text}")
    for i, opt in enumerate(options, 1):
        marker = " (*)" if opt["value"] == default else ""
        print(f"    {i}. {opt['label']}{marker}")
    while True:
        raw = input(f"  Elige [1-{len(options)}]{f' (default: {default})' if default else ''}: ").strip()
        if not raw and default is not None:
            return default
        try:
            idx = int(raw)
            if 1 <= idx <= len(options):
                return options[idx - 1]["value"]
        except ValueError:
            pass
        print(f"    Opcion invalida. Ingresa un numero entre 1 y {len(options)}.")


def _ask_text(prompt_text, default="", required=False, multiline=False):
    """Ask for free text input."""
    req = " (requerido)" if required else " (opcional, Enter para omitir)"
    print(f"\n  {prompt_text}{req}")
    if multiline:
        print("  (Escribe tu texto. Linea vacia para terminar)")
        lines = []
        while True:
            line = input("  > ")
            if not line:
                break
            lines.append(line)
        result = " ".join(lines).strip()
    else:
        result = input(f"  > ").strip()

    if not result and default:
        return default
    if not result and required:
        print("    Este campo es requerido.")
        return _ask_text(prompt_text, default, required, multiline)
    return result


def _ask_list(prompt_text, separator=","):
    """Ask for a comma-separated list."""
    print(f"\n  {prompt_text} (opcional, separados por coma)")
    raw = input("  > ").strip()
    if not raw:
        return []
    return [item.strip() for item in raw.split(separator) if item.strip()]


def run_interactive_wizard():
    """Step-by-step wizard to build a campaign config."""
    print()
    print("=" * 60)
    print("  SKYDROPX AD GENERATOR - Asistente de campana")
    print("=" * 60)
    print()
    print("  Te voy a guiar paso a paso para configurar tu campana.")
    print("  En cada paso te explico que significa el campo.")
    print()

    # --- Step 1: Canal ---
    print("─" * 60)
    print("  PASO 1/10: CANAL")
    print("─" * 60)
    print("  Selecciona la plataforma donde se publicaran los anuncios.")
    print("  Cada canal tiene formatos y reglas diferentes.")
    canal = _ask_choice(
        "En que canal vas a publicar?",
        [
            {"value": "google_ads", "label": "Google Ads   - Search RSA, Performance Max"},
            {"value": "meta", "label": "Meta Ads     - Imagen, Video, Carousel (Facebook/Instagram)"},
        ],
        default="google_ads",
    )

    # --- Step 2: Formato (depende del canal) ---
    print()
    print("─" * 60)
    print("  PASO 2/10: FORMATO DE ANUNCIO")
    print("─" * 60)

    if canal == "google_ads":
        print("  Define el tipo de anuncio de Google Ads que vas a generar.")
        print("  Cada formato tiene sus propias reglas de caracteres.")
        formato = _ask_choice(
            "Que formato de anuncio necesitas?",
            [
                {"value": "search_rsa", "label": "Search RSA       - 15 titulares (30ch) + 4 descripciones (90ch)"},
                {"value": "pmax", "label": "Performance Max  - 5 short headlines + 5 long headlines + 5 desc"},
            ],
            default="search_rsa",
        )
    else:
        print("  Define el tipo de anuncio de Meta Ads.")
        print("  Se generan 6 variantes (V1-V6) por campana.")
        formato = _ask_choice(
            "Que formato de anuncio necesitas?",
            [
                {"value": "meta_imagen", "label": "Imagen (Feed/Stories) - Post copy + Copy imagen + Encabezado + Desc"},
                {"value": "meta_video", "label": "Video               - Post copy + Guion de video + Encabezado + Desc"},
                {"value": "meta_carousel", "label": "Carousel            - Post copy + 3 cards + Encabezado + Desc"},
            ],
            default="meta_imagen",
        )

    # --- Step 3: Objetivo ---
    print()
    print("─" * 60)
    print("  PASO 3/10: OBJETIVO DE CAMPANA")
    print("─" * 60)
    print("  El objetivo ajusta la agresividad de los CTAs y el tono.")
    print("  - Conversion: CTAs directos ('Crea tu cuenta!', 'Agenda demo!')")
    print("  - Awareness: CTAs suaves ('Conoce Skydropx!', 'Descubre mas!')")
    print("  - Consideration: CTAs exploratorios ('Compara opciones!')")
    objetivo = _ask_choice(
        "Cual es el objetivo de la campana?",
        [
            {"value": "conversion", "label": "Conversion    - Impulsar accion inmediata"},
            {"value": "awareness", "label": "Awareness     - Posicionar marca y reconocimiento"},
            {"value": "consideration", "label": "Consideration - Educar y generar interes"},
        ],
        default="conversion",
    )

    # --- Step 4: Tier ---
    print()
    print("─" * 60)
    print("  PASO 4/10: TIER / AUDIENCIA")
    print("─" * 60)
    print("  El tier define a quien le hablas. Cambia el tono y mensajes:")
    print("  - Tier 1: Empresas grandes. Tono: datos, KPIs, automatizacion.")
    print("  - Tier 2: Medianos. Tono: crecer, agilidad, diversidad.")
    print("  - Tier 3: Emprendedores. Tono: cercano, empatico, escalable.")
    tier = _ask_choice(
        "A que audiencia va dirigida la campana?",
        [
            {"value": 1, "label": "Tier 1 - Grandes ecommerce (35-45 anos, equipo de 7+, alto volumen)"},
            {"value": 2, "label": "Tier 2 - Ecommerce medianos (4 anos, equipo de 5, redes + marketplaces)"},
            {"value": 3, "label": "Tier 3 - Emprendedor digital (negocio parcial, 1-3 personas, Instagram/WhatsApp)"},
        ],
        default=2,
    )

    # --- Step 5: Brief ---
    print()
    print("─" * 60)
    print("  PASO 5/10: BRIEF DE CAMPANA")
    print("─" * 60)
    print("  Describe el concepto de la campana, que comunicar, que evitar,")
    print("  contexto (temporada, promo, etc). Es el input principal para la IA.")
    print("  Ejemplo: 'Concepto Crecer te pone a prueba. Destacar la tension")
    print("  del crecimiento. Evitar cliches del Dia de las Madres.'")
    brief = _ask_text(
        "Escribe el brief de la campana:",
        multiline=True,
    )

    # --- Step 6: Mensajes clave ---
    print()
    print("─" * 60)
    print("  PASO 6/10: MENSAJES CLAVE")
    print("─" * 60)
    print("  Son los 2-5 mensajes principales que DEBEN reflejarse en los copys.")
    print("  La IA los usa como ancla para generar los textos.")
    print("  Ejemplo: 'Centraliza tus envios y decide con informacion clara.'")
    mensajes_clave = []
    print("\n  Ingresa los mensajes clave uno por uno (Enter vacio para terminar):")
    idx = 1
    while True:
        msg = input(f"  Mensaje {idx}: ").strip()
        if not msg:
            break
        mensajes_clave.append(msg)
        idx += 1

    # --- Step 7: Territorios ---
    print()
    print("─" * 60)
    print("  PASO 7/10: TERRITORIOS CREATIVOS")
    print("─" * 60)
    print("  Son los angulos o ejes creativos de la campana.")
    print("  Si defines varios, se genera un SET COMPLETO de copys por cada uno.")
    print("  Ejemplo: 'Decidir todo el tiempo, Crecer es emocionante'")
    territorios = _ask_list(
        "Ingresa los territorios creativos (separados por coma):",
    )

    # --- Step 8: Enfoque narrativo ---
    print()
    print("─" * 60)
    print("  PASO 8/10: ENFOQUE NARRATIVO")
    print("─" * 60)
    print("  Define desde donde se cuenta la historia:")
    print("  - Plataforma: muestra la herramienta ('cotiza', 'rastrea', 'integra')")
    print("  - Personaje: habla desde el usuario ('Tu negocio merece...')")
    print("  - Plataforma+Personaje: combina ambos enfoques")
    print("  - N/A: sin enfoque especifico")
    enfoque = _ask_choice(
        "Que enfoque narrativo usar?",
        [
            {"value": "plataforma", "label": "Plataforma          - Funcionalidades concretas de la herramienta"},
            {"value": "personaje", "label": "Personaje           - Perspectiva emocional del usuario"},
            {"value": "plataforma+personaje", "label": "Plataforma+Personaje - Combina funcionalidad con emocion"},
            {"value": "n/a", "label": "N/A                 - Sin enfoque especifico"},
        ],
        default="n/a",
    )

    # --- Step 9: Pilar ---
    print()
    print("─" * 60)
    print("  PASO 9/10: PILAR DE COMUNICACION")
    print("─" * 60)
    print("  El pilar del brand book que guia los puntos a cubrir:")
    print("  - Caracteristicas: funcionalidades de la plataforma")
    print("  - Beneficios: ventajas para el usuario (ahorro, control, tiempo)")
    print("  - Marca: posicionamiento y personalidad de Skydropx")
    pilar = _ask_choice(
        "Que pilar de comunicacion?",
        [
            {"value": "caracteristicas", "label": "Caracteristicas - Funcionalidades de la plataforma"},
            {"value": "beneficios", "label": "Beneficios      - Ventajas para el usuario"},
            {"value": "marca", "label": "Marca           - Posicionamiento de Skydropx"},
        ],
        default="caracteristicas",
    )

    # --- Step 10: Nota ---
    print()
    print("─" * 60)
    print("  PASO 10/10: NOTA ADICIONAL")
    print("─" * 60)
    print("  Cualquier instruccion extra, restriccion o contexto.")
    print("  Ejemplo: 'Evitar cliches del Dia de las Madres'")
    nota = _ask_text("Nota adicional:")

    # Build campaign dict
    campaign = {
        "canal": canal,
        "formato": formato,
        "objetivo": objetivo,
        "tier": tier,
        "pilar": pilar,
        "brief": brief,
        "mensajes_clave": mensajes_clave,
        "territorios": territorios,
        "enfoque_narrativo": enfoque,
        "nota": nota,
    }

    # Confirmation
    print()
    print("=" * 60)
    print("  RESUMEN DE CAMPANA")
    print("=" * 60)
    _print_campaign_summary(campaign)

    confirm = input("\n  Generar copys con esta configuracion? [S/n]: ").strip().lower()
    if confirm and confirm not in ("s", "si", "y", "yes"):
        print("  Cancelado.")
        sys.exit(0)

    return campaign


def _print_campaign_summary(campaign):
    """Print formatted campaign summary."""
    canal_label = "Google Ads" if campaign["canal"] == "google_ads" else "Meta Ads"
    print(f"  Canal:      {canal_label}")
    print(f"  Formato:    {campaign['formato'].upper().replace('_', ' ')}")
    print(f"  Objetivo:   {campaign['objetivo'].capitalize()}")
    print(f"  Tier:       {campaign['tier']} ({STYLE_GUIDE['tiers'][campaign['tier']]['name']})")
    print(f"  Pilar:      {campaign.get('pilar', 'caracteristicas')}")
    print(f"  Enfoque:    {campaign.get('enfoque_narrativo', 'n/a')}")
    if campaign.get("brief"):
        brief_preview = campaign["brief"][:80]
        print(f"  Brief:      {brief_preview}{'...' if len(campaign['brief']) > 80 else ''}")
    if campaign.get("mensajes_clave"):
        print(f"  Mensajes:   {len(campaign['mensajes_clave'])} mensajes clave")
        for i, msg in enumerate(campaign["mensajes_clave"], 1):
            print(f"              {i}. {msg[:70]}{'...' if len(msg) > 70 else ''}")
    if campaign.get("territorios"):
        terrs = campaign["territorios"]
        if terrs and isinstance(terrs[0], dict):
            labels = [f"{t['nombre']} ({t.get('enfoque', 'n/a')})" for t in terrs]
            print(f"  Territorios: {', '.join(labels)}")
        else:
            print(f"  Territorios: {', '.join(terrs)}")
    if campaign.get("nota"):
        print(f"  Nota:       {campaign['nota'][:60]}")
    if is_meta_format(campaign["formato"]):
        print(f"  Variantes:  6 (V1-V6)")


def resolve_ad_counts(formato, cantidad_override=None):
    """Resolve final ad counts: use override if provided, else defaults from AD_SPECS."""
    specs = AD_SPECS[formato]
    if not cantidad_override:
        return {field: spec["count"] for field, spec in specs.items()}

    counts = {}
    for field, spec in specs.items():
        counts[field] = cantidad_override.get(field, spec["count"])
    return counts


# ---------------------------------------------------------------------------
# Prompt building
# ---------------------------------------------------------------------------

def build_generation_prompt(campaign, territorio=""):
    """Build the full LLM prompt from a campaign config dict."""
    formato = campaign["formato"]
    tier = campaign["tier"]
    pilar = campaign.get("pilar", "caracteristicas")
    objetivo = campaign["objetivo"]
    enfoque = campaign.get("enfoque_narrativo", "n/a")
    brief = campaign.get("brief", "")
    mensajes_clave = campaign.get("mensajes_clave", [])
    nota = campaign.get("nota", "")

    specs = AD_SPECS[formato]
    counts = resolve_ad_counts(formato, campaign.get("cantidad"))
    style_guide = build_style_guide_prompt(tier, "skydropx", pilar)
    pilar_data = PILAR_CONTENT.get("skydropx", {}).get(pilar, {})

    if is_meta_format(formato):
        return _build_meta_prompt(
            formato, specs, counts, style_guide, pilar_data,
            tier, pilar, objetivo, enfoque, brief, mensajes_clave, territorio, nota,
        )
    else:
        return _build_google_ads_prompt(
            formato, specs, counts, style_guide, pilar_data,
            tier, pilar, objetivo, enfoque, brief, mensajes_clave, territorio, nota,
        )


def _build_google_ads_prompt(formato, specs, counts, style_guide, pilar_data,
                              tier, pilar, objetivo, enfoque, brief, mensajes_clave,
                              territorio, nota):
    """Build prompt for Google Ads formats."""
    prompt_parts = [
        "Eres un experto copywriter de performance marketing para Google Ads.",
        "Tu trabajo es generar copys publicitarios en ESPANOL para la marca Skydropx.",
        "Debes imitar el estilo, tono y estructura de los copys reales de Skydropx que se muestran abajo.",
        "",
        style_guide,
        "",
        "=== ANALISIS DE PATRONES DE COPYS REALES ===",
        COPY_PATTERNS,
        "",
    ]

    prompt_parts.append(get_headline_exemplars_for_prompt(max_per_category=4))
    prompt_parts.append("")
    prompt_parts.append(get_description_exemplars_for_prompt(max_items=8))
    prompt_parts.append("")
    prompt_parts.append("CTAs MAS USADOS EN DESCRIPTIONS REALES:")
    for cta in CTA_EXEMPLARS[:8]:
        prompt_parts.append(f"  - {cta}")
    prompt_parts.append("")

    # Shared sections
    _append_shared_sections(prompt_parts, objetivo, enfoque, brief, mensajes_clave, territorio)

    # Task header
    prompt_parts.extend([
        f"=== TAREA: Generar copys para {formato.upper().replace('_', ' ')} ===",
        f"Marca: Skydropx",
        f"Tier: {tier} ({STYLE_GUIDE['tiers'][tier]['name']})",
        f"Pilar: {pilar}",
        "",
    ])

    _append_pilar_instructions(prompt_parts, pilar_data)

    if nota:
        prompt_parts.extend(["=== NOTA ADICIONAL ===", nota, ""])

    # Learning from feedback
    learning = get_learning_prompt_section(formato)
    if learning:
        prompt_parts.append(learning)
        prompt_parts.append("")

    # Output format
    prompt_parts.append("=== FORMATO DE SALIDA REQUERIDO ===")
    prompt_parts.append("Responde UNICAMENTE con un JSON valido con la siguiente estructura:")
    prompt_parts.append("{")

    if formato == "search_rsa":
        h_count = counts.get("headlines", specs["headlines"]["count"])
        d_count = counts.get("descriptions", specs["descriptions"]["count"])
        prompt_parts.extend([
            f'  "headlines": ["titulo1", "titulo2", ...],  // {h_count} titulares, max {specs["headlines"]["max_chars"]} chars cada uno',
            f'  "descriptions": ["desc1", "desc2", ...]  // {d_count} descripciones, max {specs["descriptions"]["max_chars"]} chars cada una',
        ])
    elif formato == "pmax":
        prompt_parts.extend([
            f'  "short_headlines": ["...", ...],  // {counts.get("short_headlines", 5)} titulares cortos, max {specs["short_headlines"]["max_chars"]} chars',
            f'  "long_headlines": ["...", ...],  // {counts.get("long_headlines", 5)} titulares largos, max {specs["long_headlines"]["max_chars"]} chars',
            f'  "descriptions": ["...", ...],  // {counts.get("descriptions", 5)} descripciones, max {specs["descriptions"]["max_chars"]} chars',
            f'  "business_name": "..."  // max {specs["business_name"]["max_chars"]} chars',
        ])

    prompt_parts.extend([
        "}",
        "",
        "REGLAS CRITICAS:",
        "1. RESPETAR LIMITES DE CARACTERES ES LO MAS IMPORTANTE.",
        "   - Cada caracter cuenta: letras, espacios, acentos (á=1, é=1, í=1, ó=1, ú=1, ñ=1), signos de puntuacion.",
        "   - Headlines: MAXIMO 30 caracteres. Apunta a 22-28 para tener margen.",
        "   - Descripciones: MAXIMO 90 caracteres. Apunta a 75-85 para tener margen.",
        "   - CUENTA los caracteres de cada copy ANTES de incluirlo. Si excede, reescribe mas corto.",
        "2. IMITA el estilo de los ejemplos reales. Usa las mismas estructuras y verbos.",
        "3. Mezcla TECNICAS de headline: accion, beneficio, pregunta, brand-first. No uses solo una.",
        "4. Usar tono amistoso, claro, casual. NUNCA apatico, serio o irreverente.",
        "5. Skydropx siempre con S mayuscula, resto en minusculas.",
        "6. Cada copy debe ser UNICO, no repetir ideas entre titulares.",
        "7. TODAS las descripciones deben terminar con un CTA con signos de exclamacion (¡...!).",
        "8. Escribir en espanol de Mexico (tuteo informal con 'tu').",
        "9. NO incluir emojis en los copys.",
        "10. Responde SOLO con el JSON, sin texto adicional.",
    ])

    return "\n".join(prompt_parts)


def _build_meta_prompt(formato, specs, counts, style_guide, pilar_data,
                        tier, pilar, objetivo, enfoque, brief, mensajes_clave,
                        territorio, nota):
    """Build prompt for Meta Ads formats."""
    prompt_parts = [
        "Eres un experto copywriter de performance marketing para Meta Ads (Facebook e Instagram).",
        "Tu trabajo es generar copys publicitarios en ESPANOL para la marca Skydropx.",
        "Debes imitar el estilo, tono y estructura de los copys reales de campanas Meta de Skydropx.",
        "",
        style_guide,
        "",
        "=== ANALISIS DE PATRONES DE COPYS REALES DE META ===",
        META_COPY_PATTERNS,
        "",
    ]

    # Meta exemplars
    prompt_parts.append(get_meta_post_copy_exemplars_for_prompt(max_items=4))
    prompt_parts.append("")
    prompt_parts.append(get_meta_encabezado_exemplars_for_prompt())
    prompt_parts.append("")
    prompt_parts.append(get_meta_descripcion_exemplars_for_prompt())
    prompt_parts.append("")

    if formato == "meta_imagen" or formato == "meta_carousel":
        prompt_parts.append(get_meta_copy_imagen_exemplars_for_prompt(max_items=5))
        prompt_parts.append("")
    elif formato == "meta_video":
        prompt_parts.append(get_meta_video_exemplars_for_prompt(max_items=1))
        prompt_parts.append("")

    prompt_parts.append("CTAs MAS USADOS EN IMAGEN META (sin emojis, sin signos de exclamacion):")
    for cta in META_CTA_EXEMPLARS[:8]:
        prompt_parts.append(f"  - {cta}")
    prompt_parts.append("")

    # Shared sections
    _append_shared_sections(prompt_parts, objetivo, enfoque, brief, mensajes_clave, territorio)

    # Task header
    formato_label = {
        "meta_imagen": "META ADS - IMAGEN (Feed/Stories)",
        "meta_video": "META ADS - VIDEO",
        "meta_carousel": "META ADS - CAROUSEL",
    }.get(formato, formato.upper())

    prompt_parts.extend([
        f"=== TAREA: Generar 6 variantes para {formato_label} ===",
        f"Marca: Skydropx",
        f"Tier: {tier} ({STYLE_GUIDE['tiers'][tier]['name']})",
        f"Pilar: {pilar}",
        "",
    ])

    _append_pilar_instructions(prompt_parts, pilar_data)

    if nota:
        prompt_parts.extend(["=== NOTA ADICIONAL ===", nota, ""])

    # Learning from feedback
    learning = get_learning_prompt_section(formato)
    if learning:
        prompt_parts.append(learning)
        prompt_parts.append("")

    # Output format
    prompt_parts.append("=== FORMATO DE SALIDA REQUERIDO ===")
    prompt_parts.append("Genera 6 VARIANTES (V1-V6). Cada variante es un anuncio completo y diferente.")
    prompt_parts.append("Responde UNICAMENTE con un JSON valido con la siguiente estructura:")
    prompt_parts.append("{")

    if formato == "meta_imagen":
        prompt_parts.extend([
            '  "post_copy": ["v1...", "v2...", ...],        // 6 post copies, 250-450 chars cada uno',
            '  "copy_imagen": ["v1...", "v2...", ...],      // 6 copys de imagen, formato: "Título: ...\\nCopy in: ...\\nCTA: ..."',
            '  "encabezado": ["v1...", "v2...", ...],       // 6 encabezados, max 25 chars',
            '  "descripcion": ["v1...", "v2...", ...]       // 6 descripciones, max 30 chars',
        ])
    elif formato == "meta_video":
        prompt_parts.extend([
            '  "post_copy": ["v1...", "v2...", ...],        // 6 post copies, 250-450 chars cada uno',
            '  "guion_video": ["v1...", "v2...", ...],      // 6 guiones de video (4 escenas cada uno, max 1600 chars)',
            '  "encabezado": ["v1...", "v2...", ...],       // 6 encabezados, max 25 chars',
            '  "descripcion": ["v1...", "v2...", ...]       // 6 descripciones, max 30 chars',
        ])
    elif formato == "meta_carousel":
        prompt_parts.extend([
            '  "post_copy": ["v1...", "v2...", ...],        // 6 post copies, 250-450 chars cada uno',
            '  "card_1_copy": ["v1...", "v2...", ...],      // 6 copys para card 1, formato: "Título: ...\\nCopy in: ...\\nCTA: ..."',
            '  "card_2_copy": ["v1...", "v2...", ...],      // 6 copys para card 2, mismo formato',
            '  "card_3_copy": ["v1...", "v2...", ...],      // 6 copys para card 3, mismo formato',
            '  "encabezado": ["v1...", "v2...", ...],       // 6 encabezados, max 25 chars',
            '  "descripcion": ["v1...", "v2...", ...]       // 6 descripciones, max 30 chars',
        ])

    prompt_parts.extend([
        "}",
        "",
        "REGLAS CRITICAS PARA META ADS:",
        "1. RESPETAR LIMITES DE CARACTERES. Cada caracter cuenta (letras, espacios, acentos, emojis).",
        "2. POST COPY (250-450 chars):",
        "   - Estructura OBLIGATORIA: Hook → Linea en blanco → Puente → 3 bullets con emoji → Linea en blanco → CTA",
        "   - Los emojis se usan SOLO como marcadores de bullet points (al inicio de cada feature).",
        "   - Emojis permitidos para bullets: ✅, 📦, 🚀, 🚛, 💻, 📊, 💰, 🔥, ⚡, 🖥️, 🛡️, 💼",
        "   - El CTA final SI lleva signos de exclamacion: ¡...!",
        "   - NUNCA emojis sueltos al inicio del post ni como decoracion aleatoria.",
        "3. COPY DE IMAGEN / CARDS (SIN emojis):",
        "   - Formato: 'Título: [texto]\\nCopy in: [texto]\\nCTA: [texto]'",
        "   - CERO emojis. Tono ligeramente mas formal que el post copy.",
        "   - CTA sin signos de exclamacion (ej: 'Agenda una demo', NO '¡Agenda una demo!').",
        "4. ENCABEZADO (max 25 chars):",
        "   - Frase corta de accion/beneficio.",
        "   - Opcionalmente UN emoji al FINAL (63% de los reales lo usan). Ej: 'Envía fácil 📦'",
        "   - NUNCA emoji al inicio ni en medio.",
        "5. DESCRIPCION (max 30 chars):",
        "   - Frase complementaria.",
        "   - Opcionalmente UN emoji al FINAL (52% de los reales lo usan). Ej: 'Ahorra con cada envío 💸'",
        "   - NUNCA emoji al inicio ni en medio.",
    ])

    if formato == "meta_video":
        prompt_parts.extend([
            "6. GUION DE VIDEO (max 1600 chars por guion):",
            "   - 4 escenas por guion.",
            "   - Cada escena: '[Visual: descripcion]\\nVoz en off: narración\\nCrédito: texto en pantalla'",
            "   - Escena 1: PROBLEMA/HOOK, Escena 2: SOLUCION, Escena 3: FEATURES, Escena 4: CIERRE (logo+CTA)",
            "   - SIN emojis en guiones.",
        ])

    prompt_parts.extend([
        "",
        "REGLAS GENERALES:",
        "- Skydropx siempre con S mayuscula, resto en minusculas.",
        "- Cada variante debe ser UNICA (diferente angulo, hook, o enfoque).",
        "- Tono amistoso, claro, casual. NUNCA apatico, serio o irreverente.",
        "- Escribir en espanol de Mexico (tuteo informal con 'tu').",
        "- Responde SOLO con el JSON, sin texto adicional.",
    ])

    return "\n".join(prompt_parts)


def _append_shared_sections(prompt_parts, objetivo, enfoque, brief, mensajes_clave, territorio):
    """Append sections common to both Google Ads and Meta prompts."""
    obj_prompt = OBJETIVO_PROMPT.get(objetivo, OBJETIVO_PROMPT["conversion"])
    prompt_parts.extend(["=== OBJETIVO DE CAMPANA ===", obj_prompt, ""])

    enf_prompt = ENFOQUE_NARRATIVO_PROMPT.get(enfoque, "")
    if enf_prompt:
        prompt_parts.extend(["=== ENFOQUE NARRATIVO ===", enf_prompt, ""])

    if brief:
        prompt_parts.extend(["=== BRIEF DE CAMPANA ===", brief, ""])

    if mensajes_clave:
        prompt_parts.append("=== MENSAJES CLAVE (deben reflejarse en los copys) ===")
        for i, msg in enumerate(mensajes_clave, 1):
            prompt_parts.append(f"  {i}. {msg}")
        prompt_parts.append("")

    if territorio:
        prompt_parts.extend([
            "=== TERRITORIO CREATIVO ===",
            f"Genera los copys bajo el angulo/territorio: \"{territorio}\"",
            "Todos los copys deben estar alineados a este territorio.",
            "",
        ])


def _append_pilar_instructions(prompt_parts, pilar_data):
    """Append pilar-specific instructions."""
    if pilar_data.get("specs"):
        prompt_parts.append(f"Instrucciones del pilar: {pilar_data['specs']}")
        if pilar_data.get("features"):
            prompt_parts.append("Puntos a cubrir:")
            for f in pilar_data["features"]:
                prompt_parts.append(f"  - {f}")
        prompt_parts.append("")


# ---------------------------------------------------------------------------
# Generation
# ---------------------------------------------------------------------------

def generate_with_gemini(prompt, formato=""):
    max_tokens = 8192 if is_meta_format(formato) else 4096
    model = genai.GenerativeModel("gemini-2.5-flash")
    response = model.generate_content(
        prompt,
        generation_config=genai.types.GenerationConfig(
            temperature=0.9,
            top_p=0.95,
            max_output_tokens=max_tokens,
        ),
    )

    text = response.text.strip()
    json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if json_match:
        text = json_match.group(1)
    elif not text.startswith("{"):
        start = text.find("{")
        end = text.rfind("}") + 1
        if start >= 0 and end > start:
            text = text[start:end]

    return json.loads(text)


def trim_to_limit(text, max_chars):
    """Trim text to fit within character limit, breaking at last space or punctuation."""
    if len(text) <= max_chars:
        return text
    trimmed = text[:max_chars]
    last_space = trimmed.rfind(" ")
    last_period = trimmed.rfind(".")
    last_excl = trimmed.rfind("!")
    cut_point = max(last_space, last_period, last_excl)
    if cut_point > max_chars * 0.6:
        trimmed = trimmed[:cut_point]
        if trimmed[-1] in (" ", ","):
            trimmed = trimmed[:-1]
        if not trimmed.endswith((".", "!", "?")):
            trimmed += "."
    return trimmed


def normalize_ads(raw, formato):
    """Normalize the raw LLM output and enforce character limits."""
    if formato == "pmax" and isinstance(raw.get("business_name"), str):
        raw["business_name"] = [raw["business_name"]]

    specs = AD_SPECS.get(formato, {})
    for field_name, field_spec in specs.items():
        max_chars = field_spec["max_chars"]
        if field_name in raw and isinstance(raw[field_name], list):
            raw[field_name] = [trim_to_limit(t, max_chars) for t in raw[field_name]]
    return raw


def generate_for_territory(campaign, territorio, retries=3):
    """Generate ads for a single territory. Returns (ads, validation) or (None, None)."""
    prompt = build_generation_prompt(campaign, territorio)
    formato = campaign["formato"]
    best_ads = None
    best_validation = None

    for attempt in range(1, retries + 1):
        label = f"[{territorio}] " if territorio else ""
        print(f"\n  {label}Intento {attempt}/{retries}...")
        try:
            raw_ads = generate_with_gemini(prompt, formato)
            ads = normalize_ads(raw_ads, formato)
            validation = validate_ad_set(formato, ads)
            print_validation_report(validation)

            if validation["valid"]:
                return ads, validation

            if best_ads is None or len(validation["warnings"]) < len(best_validation["warnings"]):
                best_ads = ads
                best_validation = validation

        except json.JSONDecodeError as e:
            print(f"  {label}Error parseando JSON: {e}")
        except Exception as e:
            print(f"  {label}Error: {e}")

    return best_ads, best_validation


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(results, campaign):
    """Export generated ads to Excel. results = {territorio: (ads, validation)}"""
    formato = campaign["formato"]

    if is_meta_format(formato):
        return _export_meta_excel(results, campaign)
    else:
        return _export_google_ads_excel(results, campaign)


def _get_styles():
    """Return common Excel styles."""
    return {
        "header_fill": PatternFill(start_color="5233EA", end_color="5233EA", fill_type="solid"),
        "header_font": Font(name="Inter", bold=True, color="FFFFFF", size=11),
        "cell_font": Font(name="Inter", size=10),
        "bold_font": Font(name="Inter", bold=True, size=10),
        "ok_fill": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "err_fill": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "wrap_align": Alignment(wrap_text=True, vertical="top"),
    }


def _write_metadata(ws, campaign, territorio, styles):
    """Write campaign metadata to worksheet. Returns next row."""
    formato = campaign["formato"]
    canal_label = "Google Ads" if campaign["canal"] == "google_ads" else "Meta Ads"
    meta = [
        ("Generado por", "Skydropx Ad Generator"),
        ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Canal", canal_label),
        ("Formato", formato.upper().replace("_", " ")),
        ("Objetivo", campaign["objetivo"].capitalize()),
        ("Tier", f"{campaign['tier']} - {STYLE_GUIDE['tiers'][campaign['tier']]['name']}"),
        ("Enfoque", campaign.get("enfoque_narrativo", "n/a").capitalize()),
    ]
    if territorio:
        meta.append(("Territorio", territorio))
        # Show per-territory enfoque if available
        terrs = campaign.get("territorios", [])
        for t in terrs:
            if isinstance(t, dict) and t.get("nombre") == territorio:
                meta.append(("Enfoque territorio", t.get("enfoque", "n/a").capitalize()))
                break
    if campaign.get("nota"):
        meta.append(("Nota", campaign["nota"][:80]))

    for i, (label, value) in enumerate(meta, 1):
        ws.cell(row=i, column=1, value=label).font = styles["bold_font"]
        ws.cell(row=i, column=2, value=value).font = styles["cell_font"]

    row = len(meta) + 2
    if campaign.get("brief"):
        ws.cell(row=row, column=1, value="Brief").font = styles["bold_font"]
        ws.cell(row=row, column=2, value=campaign["brief"][:200]).font = styles["cell_font"]
        row += 1
    if campaign.get("mensajes_clave"):
        ws.cell(row=row, column=1, value="Mensajes clave").font = styles["bold_font"]
        for j, msg in enumerate(campaign["mensajes_clave"]):
            ws.cell(row=row, column=2 + j, value=msg).font = styles["cell_font"]
        row += 1

    return row + 1


def _export_google_ads_excel(results, campaign):
    """Export Google Ads results to Excel."""
    wb = openpyxl.Workbook()
    styles = _get_styles()
    formato = campaign["formato"]
    first_sheet = True

    for territorio, (ads, validation) in results.items():
        if ads is None:
            continue

        sheet_title = territorio[:31] if territorio else formato
        if first_sheet:
            ws = wb.active
            ws.title = sheet_title
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_title)

        row = _write_metadata(ws, campaign, territorio, styles)

        if formato == "search_rsa":
            row = _write_field_table(ws, row, "Titular", ads.get("headlines", []), 30, styles)
            row += 1
            row = _write_field_table(ws, row, "Descripcion", ads.get("descriptions", []), 90, styles)
        else:
            for field_name, texts in ads.items():
                if not isinstance(texts, list):
                    texts = [texts]
                field_spec = AD_SPECS[formato].get(field_name, {})
                max_chars = field_spec.get("max_chars", 90)
                label = field_name.replace("_", " ").title()
                row = _write_field_table(ws, row, label, texts, max_chars, styles)
                row += 1

        if validation and validation.get("warnings"):
            row += 1
            ws.cell(row=row, column=1, value="Advertencias").font = styles["header_font"]
            ws.cell(row=row, column=1).fill = styles["header_fill"]
            ws.cell(row=row, column=2).fill = styles["header_fill"]
            row += 1
            for w in validation["warnings"]:
                ws.cell(row=row, column=1, value=w).font = styles["cell_font"]
                row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 65
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 10

    return _save_workbook(wb, campaign, results)


def _export_meta_excel(results, campaign):
    """Export Meta Ads results to Excel with variant-based layout."""
    wb = openpyxl.Workbook()
    styles = _get_styles()
    formato = campaign["formato"]
    first_sheet = True

    # Define field display config per format
    field_config = _get_meta_field_config(formato)

    for territorio, (ads, validation) in results.items():
        if ads is None:
            continue

        sheet_title = territorio[:31] if territorio else formato.replace("meta_", "")
        if first_sheet:
            ws = wb.active
            ws.title = sheet_title
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_title)

        row = _write_metadata(ws, campaign, territorio, styles)
        row += 1

        # Header row for variant table
        headers = ["Variante"] + [fc["label"] for fc in field_config] + ["Estado"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
        row += 1

        # Write each variant as a row
        for v_idx in range(6):
            variant_label = f"V{v_idx + 1}"
            ws.cell(row=row, column=1, value=variant_label).font = styles["bold_font"]

            all_ok = True
            for col_offset, fc in enumerate(field_config):
                field_name = fc["field"]
                max_chars = fc["max_chars"]
                texts = ads.get(field_name, [])
                text = texts[v_idx] if v_idx < len(texts) else ""

                cell = ws.cell(row=row, column=col_offset + 2, value=text)
                cell.font = styles["cell_font"]
                cell.alignment = styles["wrap_align"]

                if len(text) > max_chars:
                    all_ok = False

            status = "OK" if all_ok else "REVISAR"
            status_cell = ws.cell(row=row, column=len(field_config) + 2, value=status)
            status_cell.font = styles["cell_font"]
            status_cell.fill = styles["ok_fill"] if all_ok else styles["err_fill"]

            for c in range(1, len(field_config) + 3):
                ws.cell(row=row, column=c).border = styles["thin_border"]

            row += 1

        # Character count detail table
        row += 2
        ws.cell(row=row, column=1, value="Detalle de caracteres").font = styles["header_font"]
        ws.cell(row=row, column=1).fill = styles["header_fill"]
        for c in range(2, 5):
            ws.cell(row=row, column=c).fill = styles["header_fill"]
        row += 1

        detail_headers = ["Campo", "Variante", "Chars", "Max", "Estado"]
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
        row += 1

        for fc in field_config:
            field_name = fc["field"]
            max_chars = fc["max_chars"]
            texts = ads.get(field_name, [])
            for v_idx, text in enumerate(texts):
                char_count = len(text)
                valid = char_count <= max_chars
                ws.cell(row=row, column=1, value=fc["label"]).font = styles["cell_font"]
                ws.cell(row=row, column=2, value=f"V{v_idx + 1}").font = styles["cell_font"]
                ws.cell(row=row, column=3, value=char_count).font = styles["cell_font"]
                ws.cell(row=row, column=4, value=max_chars).font = styles["cell_font"]
                ws.cell(row=row, column=5, value="OK" if valid else "EXCEDE").font = styles["cell_font"]
                ws.cell(row=row, column=5).fill = styles["ok_fill"] if valid else styles["err_fill"]
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = styles["thin_border"]
                row += 1

        # Warnings
        if validation and validation.get("warnings"):
            row += 1
            ws.cell(row=row, column=1, value="Advertencias").font = styles["header_font"]
            ws.cell(row=row, column=1).fill = styles["header_fill"]
            row += 1
            for w in validation["warnings"]:
                ws.cell(row=row, column=1, value=w).font = styles["cell_font"]
                row += 1

        # Column widths
        ws.column_dimensions["A"].width = 14
        col_letter = "B"
        for fc in field_config:
            ws.column_dimensions[col_letter].width = fc.get("width", 40)
            col_letter = chr(ord(col_letter) + 1)
        ws.column_dimensions[col_letter].width = 10

    return _save_workbook(wb, campaign, results)


def _get_meta_field_config(formato):
    """Return field display configuration for Meta formats."""
    if formato == "meta_imagen":
        return [
            {"field": "post_copy", "label": "Post Copy", "max_chars": 450, "width": 60},
            {"field": "copy_imagen", "label": "Copy de la Imagen", "max_chars": 280, "width": 45},
            {"field": "encabezado", "label": "Encabezado", "max_chars": 25, "width": 25},
            {"field": "descripcion", "label": "Descripción", "max_chars": 30, "width": 28},
        ]
    elif formato == "meta_video":
        return [
            {"field": "post_copy", "label": "Post Copy", "max_chars": 450, "width": 60},
            {"field": "guion_video", "label": "Guión de Video", "max_chars": 1600, "width": 70},
            {"field": "encabezado", "label": "Encabezado", "max_chars": 25, "width": 25},
            {"field": "descripcion", "label": "Descripción", "max_chars": 30, "width": 28},
        ]
    elif formato == "meta_carousel":
        return [
            {"field": "post_copy", "label": "Post Copy", "max_chars": 450, "width": 60},
            {"field": "card_1_copy", "label": "Card 1", "max_chars": 280, "width": 40},
            {"field": "card_2_copy", "label": "Card 2", "max_chars": 280, "width": 40},
            {"field": "card_3_copy", "label": "Card 3", "max_chars": 280, "width": 40},
            {"field": "encabezado", "label": "Encabezado", "max_chars": 25, "width": 25},
            {"field": "descripcion", "label": "Descripción", "max_chars": 30, "width": 28},
        ]
    return []


def _save_workbook(wb, campaign, results):
    """Save workbook and return filepath."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    formato = campaign["formato"]
    territorios_label = f"_{len(results)}terr" if len(results) > 1 else ""
    filename = (
        f"{formato}_{campaign['objetivo']}_tier{campaign['tier']}"
        f"{territorios_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath


def _write_field_table(ws, row, label, texts, max_chars, styles):
    """Write a table of ad texts with validation columns. Returns next row."""
    headers = [("No.", 1), (label, 2), ("Chars", 3), ("Max", 4), ("Estado", 5)]
    for name, col in headers:
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
    row += 1

    for i, text in enumerate(texts, 1):
        char_count = len(text)
        valid = char_count <= max_chars
        ws.cell(row=row, column=1, value=i).font = styles["cell_font"]
        ws.cell(row=row, column=2, value=text).font = styles["cell_font"]
        ws.cell(row=row, column=3, value=char_count).font = styles["cell_font"]
        ws.cell(row=row, column=4, value=max_chars).font = styles["cell_font"]
        ws.cell(row=row, column=5, value="OK" if valid else "EXCEDE").font = styles["cell_font"]
        ws.cell(row=row, column=5).fill = styles["ok_fill"] if valid else styles["err_fill"]
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = styles["thin_border"]
        row += 1

    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    campaign = run_interactive_wizard()

    configure_gemini()

    # Generate per territory (or once if no territories)
    territorios = campaign.get("territorios", [])
    if not territorios:
        territorios = [""]

    results = {}
    for territorio in territorios:
        label = territorio if territorio else campaign["formato"]
        print(f"\n{'─' * 50}")
        print(f"  Generando: {label}")
        print(f"{'─' * 50}")

        ads, validation = generate_for_territory(campaign, territorio)
        results[territorio or campaign["formato"]] = (ads, validation)

    # Check if we got any results
    valid_results = {k: v for k, v in results.items() if v[0] is not None}
    if not valid_results:
        print("\nNo se pudieron generar copys validos para ningun territorio.")
        sys.exit(1)

    # Export
    filepath = export_to_excel(valid_results, campaign)
    print(f"\nArchivo generado: {filepath}")
    print(f"\nResumen: {len(valid_results)}/{len(results)} territorios generados exitosamente.")


if __name__ == "__main__":
    main()
